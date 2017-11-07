#!/usr/bin/env python3
import os
import openpyxl

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#dir = os.getcwd()
#print(dir)

# Open workbook
wb = load_workbook(filename = 'BomList.xlsx')   # similar to wb = openpyxl.load_workbook('example.xlsx')
ws = wb['BOM1']

# Check 
for row in range(2,300):
    col = 2 
    #_ = ws.cell(column=col, row = row, value=)
    part = ws.cell(row = row, column = col).value
    if _ is not None:
        print(part)
        

    