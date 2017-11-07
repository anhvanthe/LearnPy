#!/usr/bin/env python3
import os
import openpyxl
import json

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import urllib.request
import urllib.parse


#dir = os.getcwd()
#print(dir)

# Open workbook
wb = load_workbook(filename = 'BomList.xlsx')   # similar to wb = openpyxl.load_workbook('example.xlsx')
ws = wb['BOM1']

# Check 
for row in range(2,300):
    col = 2 
    #_ = ws.cell(column=col, row = row, value=)
    part = ws.cell(row = row, column = col).value
    if _ is not None:
        print(part)
        


import http.client

conn = http.client.HTTPSConnection("api.digikey.com")

payload = "{\"SearchOptions\":[\"RoHSCompliant\"],\"Keywords\":\"TMK107BJ104KA-T\",\"RecordCount\":\"10\",\"RecordStartPosition\":\"0\",\"Filters\":{\"CategoryIds\":[27681936],\"FamilyIds\":[85103903],\"ManufacturerIds\":[86393538],\"ParametricFilters\":[{\"ParameterId\":93641443,\"ValueId\":\"311121225449472\"}]},\"Sort\":{\"Option\":\"SortByMinimumOrderQuantity\",\"Direction\":\"Descending\",\"SortParameterId\":69132357},\"RequestedQuantity\":49}"

headers = {
    'x-ibm-client-id': "85ab7852-8280-42c8-8a9c-2302c1cb029f",
    'content-type': "application/json",
    'accept': "application/json",
    'x-digikey-locale-site': "US",
    'x-digikey-locale-language': "en",
    'x-digikey-locale-currency': "USD",
    'x-digikey-locale-shiptocountry': "US",
    'x-digikey-customer-id': "tuongpv",
    'x-digikey-partner-id': "tuongpv",
    'authorization': "DsAokJNrwP0ELpoeaONMoPMX3OlK"
    }

conn.request("POST", "/services/partsearch/v2/keywordsearch", payload, headers)

res = conn.getresponse()
data = res.read()

print(data.decode("utf-8"))