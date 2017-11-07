#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

#create a workbook
wb = Workbook()

dest_filename = 'empty_book.xlsx'

# Sheets are given a name automatically when they are created. 
# They are numbered in sequence (Sheet, Sheet1, Sheet2, ...).
# You can change this name at any time with the title property:
ws1 = wb.active
ws1.title = "range names"	# sheet title

for row in range(1,40):
	ws1.append(range(600))


# Create sheet 2 with title is Pi
ws2 = wb.create_sheet(title="Pi")	#sheet title

# Write to ws2 (Pi) at col = F, row = 5, value 3.14
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data") #Sheet title
for row in range(10, 20):
	for col in range(27, 54):
		_ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
		

print(ws3['AA10'].value)
wb.save(filename = dest_filename)